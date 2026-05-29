import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

WB_PATH = Path(__file__).parent / "test_cases_phase1.xlsx"

wb = openpyxl.load_workbook(WB_PATH)

if "Claim" in wb.sheetnames:
    del wb["Claim"]

attendance_idx = wb.sheetnames.index("Attendance")
ws = wb.create_sheet("Claim", attendance_idx + 1)

# ── Column widths ────────────────────────────────────────────────────────────
for col, w in zip("ABCDEFGHIJKL", [16.8, 20.4, 8.0, 60.0, 70.0, 14.0, 22.0, 9.6, 8.0, 15.6, 10.8, 40.0]):
    ws.column_dimensions[col].width = w

# ── Style factories ──────────────────────────────────────────────────────────
def _border():
    s = Side(style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FONT  = Font(name="Calibri", bold=True,  size=11)
HDR_FILL  = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TTL_FONT  = Font(name="Calibri", bold=True,  size=11)
TTL_FILL  = PatternFill(fill_type="solid", fgColor="FFD9E1F2")
TTL_ALIGN = Alignment(vertical="center", wrap_text=True)

PRE_FONT  = Font(name="Calibri", bold=False, size=11)
PRE_FILL  = PatternFill(fill_type="solid", fgColor="FFEAF4FF")
PRE_ALIGN = Alignment(vertical="center", wrap_text=True)

STP_FONT  = Font(name="Calibri", bold=False, size=11)
NO_FILL   = PatternFill(fill_type=None)

def _set(cell, font, fill, align, border):
    cell.font, cell.fill, cell.alignment, cell.border = font, fill, align, border

def write_header(row):
    for c, v in enumerate(["Step","Requirement Group","REQ_ID","Step Description",
                            "Expected Result","Layer","Regulatory Reference",
                            "Priority","Status","Actual Result","Pass/Fail","Evidence"], 1):
        cell = ws.cell(row=row, column=c, value=v)
        _set(cell, HDR_FONT, HDR_FILL, HDR_ALIGN, _border())

def write_title(row, text):
    ws.cell(row=row, column=1, value=text)
    _set(ws.cell(row=row, column=1), TTL_FONT, TTL_FILL, TTL_ALIGN, _border())
    ws.merge_cells(f"A{row}:K{row}")
    _set(ws.cell(row=row, column=12), TTL_FONT, TTL_FILL, TTL_ALIGN, _border())

def write_precond(row, text):
    ws.cell(row=row, column=1, value=text)
    _set(ws.cell(row=row, column=1), PRE_FONT, PRE_FILL, PRE_ALIGN, _border())
    ws.merge_cells(f"A{row}:K{row}")
    _set(ws.cell(row=row, column=12), PRE_FONT, PRE_FILL, PRE_ALIGN, _border())

# wrap: cols A(1) D(4) E(5) F(6) G(7) L(12) → True; rest → None
WRAP_COLS = {1, 4, 5, 6, 7, 12}

def write_step(row, step_id, req_grp, req_id, desc, expected, layer, reg_ref, priority, status):
    vals = [step_id, req_grp, req_id, desc, expected, layer, reg_ref, priority, status, None, None, None]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        align = Alignment(vertical="top", wrap_text=(True if c in WRAP_COLS else None))
        _set(cell, STP_FONT, NO_FILL, align, _border())

# ── Test case data ───────────────────────────────────────────────────────────
REG_BILLING  = "CMS Medicaid Billing Integrity; 42 CFR Part 455"
REG_GENERAL  = "CMS Medicaid/Medicare; State Adult Day Care Licensing"
REG_HIPAA    = "HIPAA §164.312(b); CMS Medicaid/Medicare"
REG_RBAC     = "HIPAA §164.308(a)(3); CMS Medicaid/Medicare"

TCS = [
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.1 - Duplicate claim_reference_number Exhausts All Retries Returns 409 CLAIM_DUPLICATE_REFERENCE | Claim",
 "precond": "Preconditions: billing_specialist is authenticated in tenant-aaa-001 (X-User-Status: active, X-User-MFA: true). A confirmed attendance record att-seed-001 exists for participant P1 in tenant-aaa-001. Test fixture mocks uuid4() to return a fixed value so _gen_claim_ref outputs are deterministic.",
 "steps": [
  ("TC-4.1 Step 1","Claim","4.2","Insert five claim records directly into the DB for tenant-aaa-001 with reference numbers of the form MCD-{today}-AAAAAAAA that match all 5 retry slots of _gen_claim_ref when uuid4() is mocked to 'aaaaaaaa' (direct DB insert or test fixture pre-seed).","Five pre-seeded claim records exist with the targeted reference numbers before the POST call.","DB",REG_BILLING,"High","Draft"),
  ("TC-4.1 Step 2","Claim","4.2","Send POST /claims with billing_specialist headers. Body: {\"tenant_id\":\"tenant-aaa-001\",\"participant_id\":\"<P1_id>\",\"attendance_ids\":[\"<att-seed-001_id>\"],\"payer_type\":\"medicaid\",\"procedure_code\":\"T2029\",\"date_of_service_start\":\"2026-03-01\"}.","HTTP 409 Conflict. All 5 retry attempts are exhausted. Response body contains error_code=\"CLAIM_DUPLICATE_REFERENCE\".","API",REG_BILLING,"High","Draft"),
  ("TC-4.1 Step 3","Claim","4.2","Assert response HTTP status code equals 409.","HTTP 409 Conflict.","API",REG_BILLING,"High","Draft"),
  ("TC-4.1 Step 4","Claim","4.2","Assert response body detail.error_code = \"CLAIM_DUPLICATE_REFERENCE\" and message indicates no unique reference number could be generated.","detail.error_code = \"CLAIM_DUPLICATE_REFERENCE\". Message describes reference number collision.","API",REG_BILLING,"High","Draft"),
  ("TC-4.1 Step 5","Claim","4.2","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND attendance_ids LIKE \"%<att-seed-001_id>%\". Assert count = 0 (no new claim was inserted during the failed POST).","Count = 0. No new claim record created. Only the 5 seeded records remain.","DB",REG_BILLING,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.2 - Composite Duplicate participant_id + date_of_service_start + procedure_code + payer_type Returns 409 CLAIM_DUPLICATE | Claim",
 "precond": "Preconditions: A claim record already exists in tenant-aaa-001 with participant_id=P1, date_of_service_start=\"2026-03-01\", procedure_code=\"T2029\", payer_type=\"medicaid\", claim_status=\"draft\". A second confirmed attendance record att-002 exists for P1 in tenant-aaa-001. billing_specialist is authenticated.",
 "steps": [
  ("TC-4.2 Step 1","Claim","4.2","Confirm via DB query: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND participant_id=\"<P1_id>\" AND date_of_service_start=\"2026-03-01\" AND procedure_code=\"T2029\" AND payer_type=\"medicaid\". Assert count = 1.","Exactly one claim record matching the composite key exists in the DB.","DB",REG_BILLING,"High","Draft"),
  ("TC-4.2 Step 2","Claim","4.2","Send POST /claims with billing_specialist headers. Body: {\"tenant_id\":\"tenant-aaa-001\",\"participant_id\":\"<P1_id>\",\"attendance_ids\":[\"<att-002_id>\"],\"payer_type\":\"medicaid\",\"procedure_code\":\"T2029\",\"date_of_service_start\":\"2026-03-01\"}.","HTTP 409 Conflict. error_code=\"CLAIM_DUPLICATE\". No new claim is created.","API",REG_BILLING,"High","Draft"),
  ("TC-4.2 Step 3","Claim","4.2","Assert response status code equals 409.","HTTP 409 Conflict.","API",REG_BILLING,"High","Draft"),
  ("TC-4.2 Step 4","Claim","4.2","Assert response body detail.error_code = \"CLAIM_DUPLICATE\". Assert message references the duplicate combination (participant, date of service, procedure, payer).","error_code = \"CLAIM_DUPLICATE\". Message describes the composite key collision.","API",REG_BILLING,"High","Draft"),
  ("TC-4.2 Step 5","Claim","4.2","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND participant_id=\"<P1_id>\" AND date_of_service_start=\"2026-03-01\" AND procedure_code=\"T2029\" AND payer_type=\"medicaid\". Assert count still equals 1.","Count = 1. Only the original claim exists. No duplicate record was created.","DB",REG_BILLING,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.3 - Unauthorized Roles POST /claims Returns 403 RBAC_DENIED; DB Claim Count Unchanged | Claim",
 "precond": "Preconditions: Four user accounts exist in tenant-aaa-001 with roles care_coordinator, nurse_medication_aide, physician, participant_family. All have X-User-Status: active, X-User-MFA: true. A confirmed attendance record att-003 exists for participant P1. Initial claim count in tenant-aaa-001 is recorded as N.",
 "steps": [
  ("TC-4.3 Step 1","Claim","4.3","Send POST /claims with X-User-Role: care_coordinator headers for tenant-aaa-001. Body: valid claim payload referencing att-003 with required fields. Assert response status equals 403.","HTTP 403 Forbidden. error_code=\"RBAC_DENIED\". care_coordinator is not permitted to create claims.","API",REG_RBAC,"High","Draft"),
  ("TC-4.3 Step 2","Claim","4.3","Send POST /claims with X-User-Role: nurse_medication_aide headers. Body: same valid payload. Assert response status equals 403.","HTTP 403 Forbidden. error_code=\"RBAC_DENIED\". nurse_medication_aide is not permitted to create claims.","API",REG_RBAC,"High","Draft"),
  ("TC-4.3 Step 3","Claim","4.3","Send POST /claims with X-User-Role: physician headers. Body: same valid payload. Assert response status equals 403.","HTTP 403 Forbidden. error_code=\"RBAC_DENIED\". physician is not permitted to create claims.","API",REG_RBAC,"High","Draft"),
  ("TC-4.3 Step 4","Claim","4.3","Send POST /claims with X-User-Role: participant_family headers. Body: same valid payload. Assert response status equals 403.","HTTP 403 Forbidden. error_code=\"RBAC_DENIED\". participant_family is not permitted to create claims.","API",REG_RBAC,"High","Draft"),
  ("TC-4.3 Step 5","Claim","4.3","Assert all four responses have status 403 and detail.error_code=\"RBAC_DENIED\". Confirm no response returned 201.","All 4 unauthorized roles receive HTTP 403 RBAC_DENIED. Zero successful claim creations.","API",REG_RBAC,"High","Draft"),
  ("TC-4.3 Step 6","Claim","4.3","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\". Assert count still equals N.","DB claim count = N. No new claims created by any unauthorized role attempt.","DB",REG_RBAC,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.4 - PATCH Submitted Claim Non-Status Field Returns 422 CLAIM_STATUS_IMMUTABLE; Status Transition Returns 200; PATCH Paid Claim Returns 422; DB Status Unchanged After Rejected Edits | Claim",
 "precond": "Preconditions: Claim A (C_A) exists in tenant-aaa-001 with claim_status=\"submitted\" and version=1. Claim B (C_B) exists with claim_status=\"paid\" and version=2. billing_specialist is authenticated.",
 "steps": [
  ("TC-4.4 Step 1","Claim","4.4","GET /claims/<C_A_id> with billing_specialist headers. Confirm claim_status=\"submitted\" and capture version=1.","HTTP 200. claim_status=\"submitted\", version=1.","API",REG_BILLING,"High","Draft"),
  ("TC-4.4 Step 2","Claim","4.4","PATCH /claims/<C_A_id> with body {\"version\":1,\"rejection_reason\":\"Test edit on submitted claim\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_STATUS_IMMUTABLE\". Non-status fields cannot be modified on a submitted claim.","API",REG_BILLING,"High","Draft"),
  ("TC-4.4 Step 3","Claim","4.4","Query DB: SELECT claim_status, version FROM claim WHERE claim_id=\"<C_A_id>\". Assert claim_status=\"submitted\" and version=1 (unchanged after the rejected PATCH).","claim_status=\"submitted\", version=1. No changes applied by the rejected request.","DB",REG_BILLING,"High","Draft"),
  ("TC-4.4 Step 4","Claim","4.4","PATCH /claims/<C_A_id> with body {\"version\":1,\"claim_status\":\"accepted\"}. Assert response status equals 200. Assert response body claim_status=\"accepted\" and version=2.","HTTP 200 OK. claim_status=\"accepted\", version=2. Status transition from submitted to accepted succeeds.","API",REG_BILLING,"High","Draft"),
  ("TC-4.4 Step 5","Claim","4.4","GET /claims/<C_B_id> to confirm claim_status=\"paid\" and version=2.","HTTP 200. claim_status=\"paid\", version=2.","API",REG_BILLING,"High","Draft"),
  ("TC-4.4 Step 6","Claim","4.4","PATCH /claims/<C_B_id> with body {\"version\":2,\"rejection_reason\":\"Attempt to modify paid claim\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_STATUS_IMMUTABLE\". Paid claim is fully immutable.","API",REG_BILLING,"High","Draft"),
  ("TC-4.4 Step 7","Claim","4.4","Query DB: SELECT claim_status, version FROM claim WHERE claim_id=\"<C_B_id>\". Assert claim_status=\"paid\" and version=2.","claim_status=\"paid\", version=2. No modifications applied to the paid claim.","DB",REG_BILLING,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.5 - POST /claims References Non-Confirmed Attendance Returns 422; Cross-Tenant Returns 422 or 404; Confirmed Attendance Returns 201 and Sets Attendance Status Billed | Claim",
 "precond": "Preconditions: Three attendance records exist in tenant-aaa-001: att-pending (status=\"pending\"), att-voided (status=\"voided\"), att-confirmed (status=\"confirmed\"). A fourth record att-other-tenant exists in tenant-bbb-002 (different tenant) with status=\"confirmed\". billing_specialist is authenticated with X-Tenant-Id: tenant-aaa-001.",
 "steps": [
  ("TC-4.5 Step 1","Claim","4.5","POST /claims with billing_specialist headers and attendance_ids=[att-pending_id] plus all required fields. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"ATTENDANCE_NOT_CONFIRMED\". Pending attendance cannot be used in a claim.","API",REG_BILLING,"High","Draft"),
  ("TC-4.5 Step 2","Claim","4.5","POST /claims with attendance_ids=[att-voided_id] plus all required fields. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"ATTENDANCE_NOT_CONFIRMED\". Voided attendance cannot be used in a claim.","API",REG_BILLING,"High","Draft"),
  ("TC-4.5 Step 3","Claim","4.5","POST /claims with attendance_ids=[att-other-tenant_id] from tenant-bbb-002, using billing_specialist headers for tenant-aaa-001. Assert response status equals 422 or 404.","HTTP 422 or 404. Cross-tenant attendance is not visible to tenant-aaa-001. Request rejected.","API",REG_BILLING,"High","Draft"),
  ("TC-4.5 Step 4","Claim","4.5","POST /claims with attendance_ids=[att-confirmed_id] and all required fields (tenant_id, participant_id, payer_type, procedure_code, date_of_service_start). Assert response status equals 201.","HTTP 201 Created. Response body contains claim_id, claim_reference_number, claim_status=\"draft\".","API",REG_BILLING,"High","Draft"),
  ("TC-4.5 Step 5","Claim","4.5","Assert response body contains a non-null claim_id, claim_status=\"draft\", and claim_reference_number matching pattern MCD-YYYYMMDD-XXXXXXXX or MCR-YYYYMMDD-XXXXXXXX.","claim_id present, claim_status=\"draft\", claim_reference_number in expected format.","Business Rules",REG_BILLING,"High","Draft"),
  ("TC-4.5 Step 6","Claim","4.5","Query DB: SELECT status FROM attendance WHERE attendance_id=\"<att-confirmed_id>\". Assert status=\"billed\".","att-confirmed status = \"billed\" after successful claim creation. Attendance auto-updated by the POST /claims endpoint.","DB",REG_BILLING,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.6 - POST /claims Multiple Confirmed Attendance Records Returns 201; DB units_billed Sum; Non-Existent Attendance UUID Returns 422 CLAIM_ATTENDANCE_NOT_FOUND | Claim",
 "precond": "Preconditions: Three confirmed attendance records exist in tenant-aaa-001 for participant P1: att-A (authorized_units_consumed=4.0), att-B (authorized_units_consumed=6.0), att-C (authorized_units_consumed=8.0). billing_specialist is authenticated.",
 "steps": [
  ("TC-4.6 Step 1","Claim","4.6","POST /claims with billing_specialist headers. Body: {\"tenant_id\":\"tenant-aaa-001\",\"participant_id\":\"<P1_id>\",\"attendance_ids\":[\"<att-A_id>\",\"<att-B_id>\",\"<att-C_id>\"],\"payer_type\":\"medicaid\",\"procedure_code\":\"T2029\",\"date_of_service_start\":\"2026-03-10\",\"units_billed\":18.0}. Assert response status equals 201.","HTTP 201 Created. Claim created referencing all three attendance records.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.6 Step 2","Claim","4.6","Assert response body attendance_ids contains att-A_id, att-B_id, and att-C_id. Assert response body units_billed = 18.0.","Response attendance_ids = [att-A_id, att-B_id, att-C_id]. units_billed = 18.0 (4.0 + 6.0 + 8.0).","Business Rules",REG_GENERAL,"Medium","Draft"),
  ("TC-4.6 Step 3","Claim","4.6","Query DB: SELECT units_billed FROM claim WHERE claim_id = {returned_claim_id}. Assert units_billed = 18.0.","DB units_billed = 18.0. Stored value matches caller-supplied sum.","DB",REG_GENERAL,"Medium","Draft"),
  ("TC-4.6 Step 4","Claim","4.6","Query DB: SELECT status FROM attendance WHERE attendance_id IN (att-A_id, att-B_id, att-C_id). Assert all three statuses = \"billed\".","All three referenced attendance records have status = \"billed\" after claim creation.","DB",REG_GENERAL,"Medium","Draft"),
  ("TC-4.6 Step 5","Claim","4.6","POST /claims with attendance_ids=[\"00000000-0000-0000-0000-000000000000\"] (non-existent UUID) and all other required fields. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_ATTENDANCE_NOT_FOUND\". Non-existent attendance UUID rejected.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.6 Step 6","Claim","4.6","Assert error_code = \"CLAIM_ATTENDANCE_NOT_FOUND\" and message contains the invalid UUID \"00000000-0000-0000-0000-000000000000\" for identification.","detail.error_code=\"CLAIM_ATTENDANCE_NOT_FOUND\". Message includes the non-existent UUID.","API",REG_GENERAL,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.7 - POST /claims Missing Required Fields participant_id, procedure_code, payer_type Each Return 400 | Claim",
 "precond": "Preconditions: billing_specialist is authenticated in tenant-aaa-001. A confirmed attendance record att-001 exists. Three separate payloads are prepared each omitting one required field.",
 "steps": [
  ("TC-4.7 Step 1","Claim","4.7","POST /claims omitting participant_id. Include all other required fields (tenant_id, attendance_ids, payer_type, procedure_code, date_of_service_start). Assert response status is 400 or 422 and response body contains the string \"participant_id\".","HTTP 400 or 422. Response text contains \"participant_id\" identifying the missing field.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.7 Step 2","Claim","4.7","POST /claims omitting procedure_code. Include all other required fields. Assert response status is 400 or 422 and response body contains the string \"procedure_code\".","HTTP 400 or 422. Response text contains \"procedure_code\".","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.7 Step 3","Claim","4.7","POST /claims omitting payer_type. Include all other required fields. Assert response status is 400 or 422 and response body contains the string \"payer_type\".","HTTP 400 or 422. Response text contains \"payer_type\".","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.7 Step 4","Claim","4.7","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND created_at > {test_start_time}. Assert count = 0.","No claims created from any of the three invalid requests. DB unchanged.","DB",REG_GENERAL,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.8 - POST /claims Produces PHI_WRITE Audit Event with All 11 Mandatory Fields; PATCH draft to submitted Produces PHI_DISCLOSE; GET /audit-logs Returns Both Events | Claim",
 "precond": "Preconditions: billing_specialist (X-User-Id: user-billing-001, X-Session-Id: sess-tc48) is authenticated in tenant-aaa-001. A confirmed attendance record att-audit-001 exists for participant P1. compliance_officer is available for audit log reads.",
 "steps": [
  ("TC-4.8 Step 1","Claim","4.8","POST /claims with billing_specialist headers (X-User-Id: user-billing-001, X-Session-Id: sess-tc48). Body: valid payload referencing att-audit-001. Assert response status equals 201. Record returned claim_id.","HTTP 201 Created. claim_id returned in response.","API",REG_HIPAA,"High","Draft"),
  ("TC-4.8 Step 2","Claim","4.8","GET /audit-logs?tenant_id=tenant-aaa-001&resource_type=Claim&resource_id={claim_id} using compliance_officer headers. Assert response status equals 200 and body is a non-empty JSON array.","HTTP 200 OK. Response is a JSON array with at least one audit entry.","API",REG_HIPAA,"High","Draft"),
  ("TC-4.8 Step 3","Claim","4.8","Assert the audit log array contains an entry with action_type=\"PHI_WRITE\" and resource_id={claim_id}. Assert all 11 mandatory fields are non-null: audit_id, timestamp, user_id, tenant_id, session_id, action_type, resource_type, resource_id, data_affected, source_ip, outcome.","PHI_WRITE entry present. All 11 mandatory fields non-null. outcome=\"SUCCESS\".","Business Rules",REG_HIPAA,"High","Draft"),
  ("TC-4.8 Step 4","Claim","4.8","Assert data_affected in the PHI_WRITE entry is an array of field name strings only (e.g. \"tenant_id\", \"procedure_code\", \"attendance_ids\"). Assert no PHI values such as names, SSNs, or DOBs appear anywhere in the audit entry.","data_affected contains field names only. No PHI values exposed in audit log.","Business Rules",REG_HIPAA,"High","Draft"),
  ("TC-4.8 Step 5","Claim","4.8","PATCH /claims/{claim_id} with billing_specialist headers. Body: {\"version\":1,\"claim_status\":\"submitted\"}. Assert response status equals 200 and response shows claim_status=\"submitted\".","HTTP 200 OK. claim_status = \"submitted\". draft-to-submitted transition triggers PHI_DISCLOSE audit event.","API",REG_HIPAA,"High","Draft"),
  ("TC-4.8 Step 6","Claim","4.8","GET /audit-logs?tenant_id=tenant-aaa-001&resource_type=Claim&resource_id={claim_id} using compliance_officer headers. Assert the response list contains an entry with action_type=\"PHI_DISCLOSE\".","At least one audit entry with action_type=\"PHI_DISCLOSE\" exists for the claim_id.","API",REG_HIPAA,"High","Draft"),
  ("TC-4.8 Step 7","Claim","4.8","Assert PHI_DISCLOSE entry data_affected = [\"claim_status\",\"submission_date\",\"claim_reference_number\"]. Assert outcome=\"SUCCESS\". Assert no raw PHI values appear in the event.","PHI_DISCLOSE data_affected = [\"claim_status\",\"submission_date\",\"claim_reference_number\"]. outcome=\"SUCCESS\". No PHI values exposed.","Business Rules",REG_HIPAA,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.9 - POST /claims with Empty attendance_ids Returns 422 CLAIM_NO_ATTENDANCE_RECORDS; DB No Claim Created; Caller-Supplied units_billed Stored as Provided | Claim",
 "precond": "Preconditions: billing_specialist is authenticated in tenant-aaa-001. A confirmed attendance record att-units-001 (authorized_units_consumed=4.0) exists for participant P1.",
 "steps": [
  ("TC-4.9 Step 1","Claim","4.9","POST /claims with body {\"tenant_id\":\"tenant-aaa-001\",\"participant_id\":\"<P1_id>\",\"attendance_ids\":[],\"payer_type\":\"medicaid\",\"procedure_code\":\"T2029\",\"date_of_service_start\":\"2026-03-15\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_NO_ATTENDANCE_RECORDS\". Empty attendance list rejected.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.9 Step 2","Claim","4.9","Assert response body detail.error_code = \"CLAIM_NO_ATTENDANCE_RECORDS\".","error_code = \"CLAIM_NO_ATTENDANCE_RECORDS\".","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.9 Step 3","Claim","4.9","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND date_of_service_start=\"2026-03-15\". Assert count = 0.","No claim created. DB unchanged.","DB",REG_GENERAL,"Medium","Draft"),
  ("TC-4.9 Step 4","Claim","4.9","POST /claims with attendance_ids=[att-units-001_id] and units_billed=999.0 (intentionally different from authorized_units_consumed=4.0) plus all required fields. Assert response status equals 201.","HTTP 201 Created. Server accepts the caller-supplied units_billed value without overriding it.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.9 Step 5","Claim","4.9","Assert response body units_billed = 999.0. Query DB: SELECT units_billed FROM claim WHERE claim_id = {returned_claim_id}. Assert units_billed = 999.0.","Response and DB both show units_billed = 999.0. Caller-supplied value persisted as-is.","DB",REG_GENERAL,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.10 - POST /claims with Phase 2 Fields secondary_payer_id, mco_id, prior_authorization_number Returns 400; DB No Claim Created in Each Case | Claim",
 "precond": "Preconditions: billing_specialist is authenticated in tenant-aaa-001. A confirmed attendance record att-p2-001 exists. Phase 2 fields (secondary_payer_id, mco_id, prior_authorization_number) are not part of the Phase 1 ClaimCreate schema.",
 "steps": [
  ("TC-4.10 Step 1","Claim","4.10","POST /claims with all required Phase 1 fields plus secondary_payer_id=\"payer-999\" in the request body. Assert response status equals 400.","HTTP 400 Bad Request. secondary_payer_id identified as a disallowed field. No claim created.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.10 Step 2","Claim","4.10","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND created_at > {test_start_time}. Assert count = 0.","No claim created by the secondary_payer_id request.","DB",REG_GENERAL,"Medium","Draft"),
  ("TC-4.10 Step 3","Claim","4.10","POST /claims with all required fields plus mco_id=\"mco-001\". Assert response status equals 400.","HTTP 400 Bad Request. mco_id is not an accepted field in Phase 1 claim schema.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.10 Step 4","Claim","4.10","Query DB: assert claim count still unchanged (no new records from step 3 request).","No claim created by the mco_id request. Count unchanged.","DB",REG_GENERAL,"Medium","Draft"),
  ("TC-4.10 Step 5","Claim","4.10","POST /claims with all required fields plus prior_authorization_number=\"PA-12345\". Assert response status equals 400.","HTTP 400 Bad Request. prior_authorization_number is not accepted in Phase 1.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.10 Step 6","Claim","4.10","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND created_at > {test_start_time}. Assert count = 0 across all three Phase 2 field requests.","No claims created by any of the three rejected requests. DB fully unchanged.","DB",REG_GENERAL,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.11 - PATCH /claims Stale Version Returns 409 CLAIM_VERSION_CONFLICT; Submitted Claim Returns 422 Before Version Check; Correct Version Returns 200 and DB Version Incremented | Claim",
 "precond": "Preconditions: Claim A (C_A) exists in tenant-aaa-001 with claim_status=\"draft\" and version=1. Claim B (C_B) exists with claim_status=\"submitted\" and version=1. billing_specialist is authenticated.",
 "steps": [
  ("TC-4.11 Step 1","Claim","4.11","GET /claims/<C_A_id> with billing_specialist headers. Confirm claim_status=\"draft\" and version=1.","HTTP 200. claim_status=\"draft\", version=1.","API",REG_BILLING,"High","Draft"),
  ("TC-4.11 Step 2","Claim","4.11","PATCH /claims/<C_A_id> with body {\"version\":0,\"rejection_reason\":\"stale version test\"}. Assert response status equals 409.","HTTP 409 Conflict. error_code=\"CLAIM_VERSION_CONFLICT\". Stale version rejected before any field changes applied.","API",REG_BILLING,"High","Draft"),
  ("TC-4.11 Step 3","Claim","4.11","Query DB: SELECT version, claim_status FROM claim WHERE claim_id=\"<C_A_id>\". Assert version=1 and claim_status=\"draft\" (unchanged).","version=1, claim_status=\"draft\". No changes applied by the rejected PATCH.","DB",REG_BILLING,"High","Draft"),
  ("TC-4.11 Step 4","Claim","4.11","PATCH /claims/<C_B_id> (claim_status=\"submitted\") with any body including correct version. Assert response status equals 422 with error_code=\"CLAIM_STATUS_IMMUTABLE\". Confirm 422 is returned before the version check is performed.","HTTP 422 before version check. CLAIM_STATUS_IMMUTABLE returned for submitted claim regardless of version correctness.","Business Rules",REG_BILLING,"High","Draft"),
  ("TC-4.11 Step 5","Claim","4.11","PATCH /claims/<C_A_id> with body {\"version\":1,\"rejection_reason\":\"corrected field value\"}. Assert response status equals 200.","HTTP 200 OK. Update accepted with correct version. Response body contains updated rejection_reason.","API",REG_BILLING,"High","Draft"),
  ("TC-4.11 Step 6","Claim","4.11","Assert response body version = 2 (incremented from 1 to 2).","version = 2 in response body. Version correctly incremented on successful PATCH.","Business Rules",REG_BILLING,"High","Draft"),
  ("TC-4.11 Step 7","Claim","4.11","Query DB: SELECT version FROM claim WHERE claim_id=\"<C_A_id>\". Assert version = 2.","DB version = 2. Persistent version increment confirmed.","DB",REG_BILLING,"High","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.12 - POST /claims Cross-Tenant Attendance Reference Returns 422 or 404; DB No Cross-Tenant Claim Created | Claim",
 "precond": "Preconditions: Attendance record att-tenant-bbb exists in tenant-bbb-002 with status=\"confirmed\". billing_specialist is authenticated with X-Tenant-Id: tenant-aaa-001.",
 "steps": [
  ("TC-4.12 Step 1","Claim","4.12","POST /claims with billing_specialist headers (X-Tenant-Id: tenant-aaa-001). Body: {\"tenant_id\":\"tenant-aaa-001\",\"participant_id\":\"<P1_id>\",\"attendance_ids\":[\"<att-tenant-bbb_id>\"],\"payer_type\":\"medicaid\",\"procedure_code\":\"T2029\",\"date_of_service_start\":\"2026-04-01\"}. Assert response status equals 422 or 404.","HTTP 422 or 404. Cross-tenant attendance not found or inaccessible from tenant-aaa-001.","API",REG_BILLING,"Medium","Draft"),
  ("TC-4.12 Step 2","Claim","4.12","Assert response body error_code is \"ATTENDANCE_NOT_FOUND\" or \"NOT_FOUND\" indicating the cross-tenant attendance was not accessible.","error_code = \"ATTENDANCE_NOT_FOUND\" or \"NOT_FOUND\". Attendance from another tenant is rejected.","API",REG_BILLING,"Medium","Draft"),
  ("TC-4.12 Step 3","Claim","4.12","Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id=\"tenant-aaa-001\" AND date_of_service_start=\"2026-04-01\". Assert count = 0.","No claim record created in tenant-aaa-001 referencing the cross-tenant attendance.","DB",REG_BILLING,"Medium","Draft"),
  ("TC-4.12 Step 4","Claim","4.12","Query DB: SELECT COUNT(*) FROM claim WHERE attendance_ids LIKE \"%<att-tenant-bbb_id>%\". Assert count = 0.","No claim in any tenant references the cross-tenant attendance in an unauthorized manner.","DB",REG_BILLING,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.13 - POST /claims Referencing Attendance Already Linked to Existing Claim Returns 422 ATTENDANCE_NOT_CONFIRMED; DB No Duplicate Claim | Claim",
 "precond": "Preconditions: Attendance record att-billed-001 exists in tenant-aaa-001 with status=\"billed\", already linked to claim C_existing. billing_specialist is authenticated.",
 "steps": [
  ("TC-4.13 Step 1","Claim","4.13","GET /attendance/<att-billed-001_id> with billing_specialist headers. Assert response status equals 200 and status=\"billed\".","HTTP 200. attendance status=\"billed\". Already associated with an existing claim.","API",REG_BILLING,"Medium","Draft"),
  ("TC-4.13 Step 2","Claim","4.13","POST /claims with body {\"tenant_id\":\"tenant-aaa-001\",\"participant_id\":\"<P2_id>\",\"attendance_ids\":[\"<att-billed-001_id>\"],\"payer_type\":\"medicaid\",\"procedure_code\":\"T2029\",\"date_of_service_start\":\"2026-05-01\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"ATTENDANCE_NOT_CONFIRMED\". Billed attendance cannot be claimed again.","API",REG_BILLING,"Medium","Draft"),
  ("TC-4.13 Step 3","Claim","4.13","Assert response body error_code = \"ATTENDANCE_NOT_CONFIRMED\" and message indicates att-billed-001 is not in confirmed status.","error_code=\"ATTENDANCE_NOT_CONFIRMED\". Message identifies att-billed-001 as already billed.","API",REG_BILLING,"Medium","Draft"),
  ("TC-4.13 Step 4","Claim","4.13","Query DB: SELECT COUNT(*) FROM claim WHERE attendance_ids LIKE \"%<att-billed-001_id>%\" AND claim_id != \"<C_existing_id>\". Assert count = 0.","Count = 0. att-billed-001 is not linked to any claim other than C_existing.","DB",REG_BILLING,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.14 - GET /claims Non-Existent claim_id Returns 404 NOT_FOUND | Claim",
 "precond": "Preconditions: billing_specialist is authenticated in tenant-aaa-001. The UUID \"00000000-0000-0000-0000-999999999999\" does not exist in the claim table.",
 "steps": [
  ("TC-4.14 Step 1","Claim","4.1","Send GET /claims/00000000-0000-0000-0000-999999999999 with billing_specialist headers (X-User-Role: billing_specialist, X-User-Status: active, X-Tenant-Id: tenant-aaa-001).","HTTP 404 Not Found. Response body contains error_code=\"NOT_FOUND\".","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.14 Step 2","Claim","4.1","Assert response status code equals 404.","HTTP 404 Not Found.","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.14 Step 3","Claim","4.1","Assert response body detail.error_code = \"NOT_FOUND\".","detail.error_code = \"NOT_FOUND\".","API",REG_GENERAL,"Medium","Draft"),
  ("TC-4.14 Step 4","Claim","4.1","Assert response body detail.message contains the string \"00000000-0000-0000-0000-999999999999\" or \"Claim\" to identify what resource was not found.","message contains the requested claim_id or the resource type \"Claim\" for traceability.","API",REG_GENERAL,"Medium","Draft"),
 ]
},
# ─────────────────────────────────────────────────────────────────────────────
{
 "title": "TC-4.15 - PATCH /claims Paid Claim Any Non-Status Field Returns 422 CLAIM_STATUS_IMMUTABLE; DB Claim Unchanged | Claim",
 "precond": "Preconditions: Claim C_paid exists in tenant-aaa-001 with claim_status=\"paid\" and version=3. billing_specialist is authenticated.",
 "steps": [
  ("TC-4.15 Step 1","Claim","4.4","GET /claims/<C_paid_id> with billing_specialist headers. Confirm claim_status=\"paid\" and capture version=3.","HTTP 200. claim_status=\"paid\", version=3.","API",REG_BILLING,"High","Draft"),
  ("TC-4.15 Step 2","Claim","4.4","PATCH /claims/<C_paid_id> with body {\"version\":3,\"rejection_reason\":\"Test non-status field edit\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_STATUS_IMMUTABLE\". Paid claim is fully immutable.","API",REG_BILLING,"High","Draft"),
  ("TC-4.15 Step 3","Claim","4.4","PATCH /claims/<C_paid_id> with body {\"version\":3,\"claim_status\":\"draft\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_STATUS_IMMUTABLE\". Status cannot be rolled back from \"paid\".","API",REG_BILLING,"High","Draft"),
  ("TC-4.15 Step 4","Claim","4.4","PATCH /claims/<C_paid_id> with body {\"version\":3,\"submission_date\":\"2026-01-01T00:00:00Z\"}. Assert response status equals 422.","HTTP 422 Unprocessable Entity. error_code=\"CLAIM_STATUS_IMMUTABLE\". All field changes rejected on paid claim.","API",REG_BILLING,"High","Draft"),
  ("TC-4.15 Step 5","Claim","4.4","Query DB: SELECT claim_status, version, rejection_reason FROM claim WHERE claim_id=\"<C_paid_id>\". Assert claim_status=\"paid\", version=3, rejection_reason is null or unchanged.","claim_status=\"paid\", version=3. All fields unchanged. No modifications persisted by any of the three rejected PATCH attempts.","DB",REG_BILLING,"High","Draft"),
 ]
},
]

# ── Write rows ───────────────────────────────────────────────────────────────
current_row = 1
write_header(current_row)
current_row += 1

for tc in TCS:
    write_title(current_row, tc["title"])
    current_row += 1
    write_precond(current_row, tc["precond"])
    current_row += 1
    for step in tc["steps"]:
        write_step(current_row, *step)
        current_row += 1

wb.save(WB_PATH)
print(f"Saved. Total rows written: {current_row - 1}. Sheets: {wb.sheetnames}")
