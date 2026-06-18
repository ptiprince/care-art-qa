"""
Fix merge structure in 'Participant Test Cases' sheet of test_cases_phase1.xlsx.
- Reads fill, font, border from R2 (title) and R3 (preconditions) as reference.
- Unmerges rows 47, 53, 54, 61, 62, 68, 69, 74, 75 (step rows).
- Adds merges A49:K49, A55:K55, A56:K56, A64:K64, A65:K65,
  A71:K71, A72:K72, A77:K77, A78:K78 with correct formatting.
- Does not alter any cell values. Does not touch other sheets.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

TARGET = "/Users/ievgeniiasorochenko/files_claude/care_art/test_cases/test_cases_phase1.xlsx"
wb = openpyxl.load_workbook(TARGET)
ws = wb["Participant Test Cases"]

# ── Read reference formatting from R2 (title) and R3 (preconditions) ──────
def read_fmt(cell):
    """Return (font, fill, border, alignment) cloned from cell."""
    f = cell.font
    p = cell.fill
    b = cell.border
    a = cell.alignment

    font = Font(
        name=f.name, bold=f.bold, italic=f.italic, size=f.sz,
        underline=f.u, strike=f.strike,
        color=f.color.rgb if f.color and f.color.type == "rgb" else None,
    )
    fill = PatternFill(
        fill_type=p.patternType,
        fgColor=p.fgColor.rgb if p.fgColor else "00000000",
        bgColor=p.bgColor.rgb if p.bgColor else "00000000",
    )
    def clone_side(s):
        return Side(border_style=s.border_style, color=s.color.rgb if s.color and s.color.type == "rgb" else None)
    border = Border(
        left=clone_side(b.left),
        right=clone_side(b.right),
        top=clone_side(b.top),
        bottom=clone_side(b.bottom),
    )
    align = Alignment(
        wrap_text=a.wrap_text,
        horizontal=a.horizontal,
        vertical=a.vertical,
    )
    return font, fill, border, align

title_fmt  = read_fmt(ws["A2"])
precond_fmt = read_fmt(ws["A3"])


def apply_fmt(cell, fmt):
    cell.font, cell.fill, cell.border, cell.alignment = fmt


# ── Unmerge step rows (already done, but idempotent) ──────────────────────
STEP_ROWS = [47, 53, 54, 61, 62, 68, 69, 74, 75]

for row in STEP_ROWS:
    to_remove = [
        str(mr) for mr in ws.merged_cells.ranges
        if mr.min_row == row and mr.max_row == row
    ]
    for rng in to_remove:
        anchor_val = ws[rng.split(":")[0]].value
        ws.unmerge_cells(rng)
        ws[rng.split(":")[0]].value = anchor_val
        print(f"  Unmerged {rng}")


# ── Add missing merges with correct formatting ─────────────────────────────
def add_merge(rng, fmt):
    # Skip if already merged
    anchor_ref = rng.split(":")[0]
    if any(str(mr) == rng for mr in ws.merged_cells.ranges):
        print(f"  Already merged {rng} — skipping")
        return
    anchor_val = ws[anchor_ref].value      # preserve value
    ws.merge_cells(rng)
    ws[anchor_ref].value = anchor_val      # restore after merge
    apply_fmt(ws[anchor_ref], fmt)
    print(f"  Merged {rng}")


# A49: preconditions row for TC-1.8; title formatting per instruction
add_merge("A49:K49", title_fmt)

# Pairs (title, preconditions) for TC-1.9 through TC-1.12
for title_row, precond_row in [(55, 56), (64, 65), (71, 72), (77, 78)]:
    add_merge(f"A{title_row}:K{title_row}", title_fmt)
    add_merge(f"A{precond_row}:K{precond_row}", precond_fmt)

wb.save(TARGET)
print(f"\nSaved: {TARGET}")
