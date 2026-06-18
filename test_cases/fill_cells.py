"""
Write specific cells into test_cases_phase1.xlsx.
Step rows in 'Participant Test Cases' that were accidentally merged A:K are
unmerged first (anchor value preserved); then all listed cells are written.
No sheets created. No structure altered beyond removing accidental step-row merges.
"""
import openpyxl

TARGET = "/Users/ievgeniiasorochenko/files_claude/care_art/test_cases/test_cases_phase1.xlsx"
wb = openpyxl.load_workbook(TARGET)


def unmerge_single_row(ws, row):
    """Remove any full-row merge on `row`, preserving the anchor cell value."""
    to_remove = [str(mr) for mr in ws.merged_cells.ranges
                 if mr.min_row == row and mr.max_row == row]
    for rng in to_remove:
        anchor_ref = rng.split(":")[0]
        anchor_val = ws[anchor_ref].value
        ws.unmerge_cells(rng)
        ws[anchor_ref].value = anchor_val


# ── Participant Test Cases ─────────────────────────────────────────────────
ws = wb["Participant Test Cases"]

# These step rows are accidentally merged A:K; unmerge before writing D–I.
for r in [47, 53, 54, 61, 62, 68, 69, 74, 75]:
    unmerge_single_row(ws, r)

ws["D47"] = "Query DB: SELECT program_status FROM participant WHERE participant_id=<participant_id>. Assert program_status='on_leave'."
ws["E47"] = "DB confirms program_status='on_leave' persisted correctly after the transition."
ws["F47"] = "DB"
ws["G47"] = "State Adult Day Care Licensing"
ws["H47"] = "Medium"
ws["I47"] = "Draft"

ws["D53"] = "Assert error_code contains 'TRANSITION'."
ws["E53"] = "HTTP 422 Unprocessable Entity. The transition from deceased to active is blocked. The error code indicates an invalid state transition. The participant record remains unchanged."
ws["F53"] = "Business Rules"
ws["G53"] = "State Adult Day Care Licensing"
ws["H53"] = "Medium"
ws["I53"] = "Draft"

ws["D54"] = "Query DB: SELECT program_status FROM participant WHERE participant_id=<participant_id>. Assert program_status='deceased' (unchanged after rejected transition)."
ws["E54"] = "DB confirms program_status remains 'deceased'. The rejected PATCH did not alter the row."
ws["F54"] = "DB"
ws["G54"] = "State Adult Day Care Licensing"
ws["H54"] = "Medium"
ws["I54"] = "Draft"

ws["D61"] = "Send GET /participants/{participant_id} using a standard role (e.g., care_coordinator)."
ws["E61"] = "HTTP 200 OK with is_deleted=true in the response. A subsequent standard GET on the same participant_id returns 404. The database row is not physically removed."
ws["F61"] = "API"
ws["G61"] = "HIPAA §164.530(j) - Record Retention"
ws["H61"] = "High"
ws["I61"] = "Draft"

ws["D62"] = "Assert response status is 404 Not Found."
ws["E62"] = "HTTP 200 OK with is_deleted=true in the response. A subsequent standard GET on the same participant_id returns 404. The database row is not physically removed."
ws["F62"] = "API"
ws["G62"] = "HIPAA §164.530(j) - Record Retention"
ws["H62"] = "High"
ws["I62"] = "Draft"

ws["D68"] = "Assert response status is 405 Method Not Allowed."
ws["E68"] = "HTTP 405 Method Not Allowed. The physical database row is not removed. The record remains retrievable by compliance_officer with include_deleted=true."
ws["F68"] = "API"
ws["G68"] = "HIPAA §164.530(j) - Record Retention"
ws["H68"] = "High"
ws["I68"] = "Draft"

ws["D69"] = "Send GET /participants/{participant_id} using compliance_officer headers with include_deleted=true."
ws["E69"] = "HTTP 405 Method Not Allowed. The physical database row is not removed. The record remains retrievable by compliance_officer with include_deleted=true."
ws["F69"] = "API"
ws["G69"] = "HIPAA §164.530(j) - Record Retention"
ws["H69"] = "High"
ws["I69"] = "Draft"

ws["D74"] = "Send POST /participants using program_administrator headers."
ws["E74"] = "HTTP 400 or 422. The error response explicitly names 'first_name' as the missing or invalid field. No participant record is created."
ws["F74"] = "API"
ws["G74"] = "CMS Medicaid/Medicare"
ws["H74"] = "High"
ws["I74"] = "Draft"

ws["D75"] = "Assert response status is 400 or 422."
ws["E75"] = "HTTP 400 or 422. The error response explicitly names 'first_name' as the missing or invalid field. No participant record is created."
ws["F75"] = "API"
ws["G75"] = "CMS Medicaid/Medicare"
ws["H75"] = "High"
ws["I75"] = "Draft"

# ── Attendance ─────────────────────────────────────────────────────────────
ws = wb["Attendance"]

ws["C16"] = "3.2"
ws["C17"] = "3.2"
ws["C18"] = "3.2"
ws["C19"] = "3.2"
ws["C22"] = "3.2"
ws["C23"] = "3.2"
ws["C24"] = "3.2"
ws["C25"] = "3.2"

ws["D54"] = "Send POST /attendance with total_hours=8 for the same Medicaid participant on fixture_date_of_service_b (a separate date offset, distinct from fixture_date_of_service_a used in Step 1, to avoid ATTENDANCE_DUPLICATE_DATE constraint)."
ws["E54"] = "HTTP 201 Created. authorized_units_consumed=32.0 (8 x 4). Server-side calculation confirmed. fixture_date_of_service_b is distinct from fixture_date_of_service_a."

# ── Claim ──────────────────────────────────────────────────────────────────
ws = wb["Claim"]

ws["D74"] = "Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id='tenant-aaa-001' AND participant_id=<P1_id>. Assert count_after equals count_before captured at test start."
ws["E74"] = "No claim created by the secondary_payer_id request. count_after equals count_before."
ws["D76"] = "SELECT COUNT(*) FROM claim WHERE tenant_id='tenant-aaa-001' AND participant_id=<P1_id>. Assert count_after equals count_before."
ws["D78"] = "Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id='tenant-aaa-001' AND participant_id=<P1_id>. Assert count_after equals count_before across all three Phase 2 field requests."
ws["E78"] = "No claims created by any of the three requests. count_after equals count_before."
ws["D93"] = "Query DB: SELECT COUNT(*) FROM claim WHERE tenant_id='tenant-aaa-001' AND participant_id=<P1_id> AND date_of_service_start=fixture_date_of_service_start. Assert count=0."
ws["E93"] = "Count=0. No claim in tenant-aaa-001 references the cross-tenant attendance."

# ── MARRecord ──────────────────────────────────────────────────────────────
ws = wb["MARRecord"]
ws["F70"] = "API"

# ── Incident ───────────────────────────────────────────────────────────────
ws = wb["Incident"]
# A72 is the anchor of A72:K72 — write directly to anchor.
ws["A72"] = "Preconditions: An incident with status='draft', version=1 exists (incident_id from fresh_incident_open fixture, which creates a draft incident). program_administrator authenticated."

# ── AuditLog ───────────────────────────────────────────────────────────────
ws = wb["AuditLog"]

ws["D4"]  = "Send POST /participants with valid payload using program_administrator headers. Record the returned participant_id."
ws["D5"]  = "Query audit_log: SELECT * FROM audit_log WHERE resource_type='Participant' AND action_type='PHI_WRITE' AND resource_id=<participant_id> ORDER BY rowid DESC LIMIT 1."
ws["D15"] = "Query audit_log: SELECT * FROM audit_log WHERE resource_type='MARRecord' AND action_type='PHI_READ' AND resource_id=<mar_id> ORDER BY rowid DESC LIMIT 1."
ws["D29"] = "Send POST /participants with valid payload using program_administrator headers. Record the returned participant_id."
ws["D30"] = "Query audit_log: SELECT * FROM audit_log WHERE resource_type='Participant' AND action_type='PHI_WRITE' AND resource_id=<participant_id> ORDER BY rowid DESC LIMIT 1."
# A37 is the anchor of A37:K37 — write directly to anchor.
ws["A37"] = "Preconditions: A program_administrator auth token is available. POST /incidents endpoint is available. Two incidents will be created in this test: one with is_sud_related=true (sud_incident_id) and one with is_sud_related=false (non_sud_incident_id). Both incident_ids are captured after their respective POST calls."
ws["D38"] = "Send POST /incidents with is_sud_related=true using program_administrator headers. Record returned incident_id as sud_incident_id. Send POST /incidents with is_sud_related=false using program_administrator headers. Record returned incident_id as non_sud_incident_id."
ws["E38"] = "Both POST calls return HTTP 201. sud_incident_id and non_sud_incident_id captured."

# ── RBACSweep ──────────────────────────────────────────────────────────────
ws = wb["RBACSweep"]

for r in [6, 11, 16, 21, 25, 29, 33, 37, 41, 45, 49, 54]:
    ws.cell(row=r, column=6).value = "DB"

ws["D49"] = "Query DB: SELECT COUNT(*) FROM participant WHERE tenant_id='tenant-aaa-001' after each POST; assert count unchanged. Repeat for user, attendance, claim, mar_record, incident tables. Assert all six counts equal count_before for each respective table."

# ── TenantIsolation ────────────────────────────────────────────────────────
ws = wb["TenantIsolation"]

ws["F5"]  = "DB"
ws["F10"] = "DB"
ws["F15"] = "DB"
ws["F20"] = "DB"
ws["F25"] = "DB"
ws["F30"] = "DB"

ws["D9"] = "Send requests to all tenant B endpoints using auth headers for tenant A (X-Tenant-Id: tenant-aaa-001)."


def insert_before_order(ws, ref, snippet):
    """Insert `snippet` immediately before 'ORDER BY rowid DESC LIMIT 1.' in cell."""
    v = ws[ref].value or ""
    marker = "ORDER BY rowid DESC LIMIT 1."
    if snippet not in v:
        if marker in v:
            ws[ref].value = v.replace(marker, snippet + " " + marker)
        else:
            ws[ref].value = (v.rstrip() + " " + snippet).strip()


insert_before_order(ws, "D6",  "AND resource_id=<participant_id_tenant_a>")
insert_before_order(ws, "D16", "AND resource_id=<attendance_id_tenant_a>")
insert_before_order(ws, "D21", "AND resource_id=<claim_id_tenant_a>")
insert_before_order(ws, "D26", "AND resource_id=<mar_id_tenant_a>")
insert_before_order(ws, "D31", "AND resource_id=<incident_id_tenant_a>")

suffix = "Use program_administrator auth headers for tenant B (X-Tenant-Id: tenant-bbb-002)."
for ref in ["D34", "D35"]:
    v = ws[ref].value or ""
    if suffix not in v:
        ws[ref].value = (v.rstrip() + " " + suffix).strip()

wb.save(TARGET)
print("Done — saved", TARGET)
